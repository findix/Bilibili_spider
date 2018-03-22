#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Spider of Bilibili
download with fiter by catelog and cluster by tag
"""

from __future__ import (absolute_import, division, print_function,
                        unicode_literals, with_statement)

from datetime import datetime
from urllib.parse import urlencode
from openpyxl import Workbook
from openpyxl import load_workbook
import requests


def get_result(page, pagesize):
    """
    pagesize must <= 100
    """
    params = {
        'main_ver': 'v3',
        'search_type': 'video',
        'view_type': 'hot_rank',
        'order': 'click',
        'copy_right': -1,
        'cate_id': 25, # 分区
        'page': page,
        'pagesize': pagesize,
        'time_from': 20171218,
        'time_to': 20180321
    }
    base_url = 'https://s.search.bilibili.com/cate/search?'
    r = requests.get(base_url, params=params)
    json = r.json()
    result = json['result']
    return result


def get_data(start_page, page_size, page_count):
    wb = Workbook()
    ws = wb.active
    ws.title = 'base'
    for page in range(start_page, start_page + page_count):
        result = get_result(page, page_size)
        for row, video_info in enumerate(result):
            fields = ['title', 'arcurl', 'author', 'duration',
                      'play', 'video_review', 'review', 'favorites', 'tag']
            for column, title_column in enumerate(fields):
                ws.cell(1, column + 1, title_column)
            for column in range(len(fields)):
                fields[column] = str(video_info[fields[column]])
                ws.cell((page - 1) * page_count + row +
                        2, column + 1, fields[column])
        print(str((page - start_page + 1) / page_count * 100) + '%')
    wb.save('bilibili_mmd.xlsx')
    print('Done')


def tag_cluster():
    tag_dict = {}
    wb = load_workbook(filename='bilibili_mmd.xlsx', guess_types=True)
    ws = wb['base']
    print("workbook loaded")
    for line_num, row in enumerate(ws.rows):
        if(line_num == 0):  # 忽略标题行
            continue
        if row[4].value == '--':
            continue
        tags = row[8].value.split(',')
        for tag in tags:
            if tag not in tag_dict:
                tag_dict[tag] = [
                    0,  # count
                    0,  # play
                    0,  # video_review
                    0,  # review
                    0   # favorites
                ]
            tag = tag_dict[tag]
            tag[0] += 1
            tag[1] += int(row[4].value)
            tag[2] += int(row[5].value)
            tag[3] += int(row[6].value)
            tag[4] += int(row[7].value)
    ws2 = wb.create_sheet('tags')
    titles = ['count', 'play', 'video_review', 'review', 'favorites']
    for column_i, column in enumerate(titles):
        ws2.cell(1, column_i + 1, column)
    for row_line, (key, values) in enumerate(tag_dict.items()):
        row_line += 2
        ws2.cell(row_line, 1, key)
        ws2.cell(row_line, 2, values[0])
        ws2.cell(row_line, 3, values[1])
        ws2.cell(row_line, 4, values[2])
        ws2.cell(row_line, 5, values[3])
        ws2.cell(row_line, 6, values[4])

    wb.save('bilibili_mmd.xlsx')
    print('Done')


def main():
    get_data(1, 10, 1)
    tag_cluster()


if __name__ == '__main__':
    main()
